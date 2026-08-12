from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_POST
from functools import wraps
from .models import ProfileAngajat, CerereConcediu, Pontaj, DocumentAngajat


@login_required
def dashboard(request):
    angajat = request.user

    profil, _ = ProfileAngajat.objects.get_or_create(
        user=angajat,
        defaults={'functia': 'Angajat', 'departament': 'General', 'data_angajare': timezone.now().date()}
    )

    if profil.rol == 'hr':
        return redirect('hr_dashboard')

    if not profil.activ:
        messages.error(request, "Contul tău a fost dezactivat. Contactează HR pentru detalii.")
        return redirect('login')

    concedii = CerereConcediu.objects.filter(angajat=angajat)
    documente = DocumentAngajat.objects.filter(angajat=angajat)

    an_curent = timezone.localtime(timezone.now()).year
    concedii_odihna_aprobate = CerereConcediu.objects.filter(
        angajat=angajat,
        tip_concediu='odihna',
        status='aprobat',
        data_inceput__year=an_curent,
    )
    zile_folosite = 0
    for c in concedii_odihna_aprobate:
        zile_calendaristice = (c.data_sfarsit - c.data_inceput).days + 1
        zile_folosite += round(zile_calendaristice * 5 / 7)
    zile_ramase = max(profil.zile_concediu_alocate - zile_folosite, 0)

    pontaje_raw = Pontaj.objects.filter(angajat=angajat).order_by('-data')
    pontaje = []
    for p in pontaje_raw:
        pontaje.append({
            'data': p.data,
            'ora_intrare': p.ora_intrare.strftime('%H:%M') if p.ora_intrare else '--',
            'ora_iesire': p.ora_iesire.strftime('%H:%M') if p.ora_iesire else '--',
            'ore_lucrate': p.ore_lucrate,
        })

    azi_local = timezone.localtime(timezone.now()).date()
    pontaje_luna = Pontaj.objects.filter(
        angajat=angajat,
        data__year=azi_local.year,
        data__month=azi_local.month,
    )
    total_ore_luna = sum(p.ore_lucrate for p in pontaje_luna)
    total_zile_lucrate = pontaje_luna.count()

    context = {
        'profil': profil,
        'concedii': concedii,
        'pontaje': pontaje,
        'documente': documente,
        'zile_folosite': zile_folosite,
        'zile_ramase': zile_ramase,
        'total_ore_luna': round(total_ore_luna, 2),
        'total_zile_lucrate': total_zile_lucrate,
    }
    return render(request, 'core/dashboard.html', context)


@login_required
def pontaj_intrare(request):
    if request.method == 'POST':
        acum_local = timezone.localtime(timezone.now())
        azi = acum_local.date()
        ora_curenta = acum_local.time()

        pontaj_azi, created = Pontaj.objects.get_or_create(
            angajat=request.user,
            data=azi,
            defaults={'ora_intrare': ora_curenta}
        )

        if not created and not pontaj_azi.ora_intrare:
            pontaj_azi.ora_intrare = ora_curenta
            pontaj_azi.save()

        messages.success(request, "Pontaj Intrare înregistrat.")

    return redirect('dashboard')


@login_required
def pontaj_iesire(request):
    if request.method == 'POST':
        acum_local = timezone.localtime(timezone.now())
        azi = acum_local.date()
        ora_curenta = acum_local.time()

        pontaj_azi, created = Pontaj.objects.get_or_create(
            angajat=request.user,
            data=azi,
            defaults={'ora_iesire': ora_curenta}
        )

        if not created:
            pontaj_azi.ora_iesire = ora_curenta
            pontaj_azi.save()

        messages.success(request, "Pontaj Ieșire înregistrat.")

    return redirect('dashboard')


def hr_required(view_func):
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        profil, _ = ProfileAngajat.objects.get_or_create(
            user=request.user,
            defaults={'functia': 'Șofer', 'departament': 'Transport', 'data_angajare': timezone.now().date()}
        )
        if profil.rol != 'hr':
            messages.error(request, "Nu ai acces la această secțiune.")
            return redirect('dashboard')
        return view_func(request, *args, **kwargs)
    return wrapper


@hr_required
def hr_dashboard(request):
    total_angajati = ProfileAngajat.objects.filter(rol='angajat').count()
    cereri_in_asteptare = CerereConcediu.objects.filter(status='in_asteptare').count()
    azi = timezone.localtime(timezone.now()).date()
    pontaje_azi = Pontaj.objects.filter(data=azi).count()
    context = {
        'total_angajati': total_angajati,
        'cereri_in_asteptare': cereri_in_asteptare,
        'pontaje_azi': pontaje_azi,
        'active_menu': 'dashboard',
    }
    return render(request, 'core/hr_dashboard.html', context)


@hr_required
def hr_angajati(request):
    query = request.GET.get('q', '').strip()
    arata_inactivi = request.GET.get('inactivi') == '1'
    profile_qs = ProfileAngajat.objects.select_related('user').filter(rol='angajat')
    if not arata_inactivi:
        profile_qs = profile_qs.filter(activ=True)
    profile_qs = profile_qs.order_by('user__username')
    if query:
        profile_qs = profile_qs.filter(
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(departament__icontains=query) |
            Q(functia__icontains=query)
        )
    paginator = Paginator(profile_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {'page_obj': page_obj, 'query': query, 'arata_inactivi': arata_inactivi, 'active_menu': 'angajati'}
    return render(request, 'core/hr_angajati.html', context)


@hr_required
def hr_pontaje(request):
    pontaje_qs = Pontaj.objects.select_related('angajat').order_by('-data')
    angajat_id = request.GET.get('angajat', '')
    if angajat_id:
        pontaje_qs = pontaje_qs.filter(angajat_id=angajat_id)
    paginator = Paginator(pontaje_qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    angajati = User.objects.filter(profile__rol='angajat').order_by('username')
    context = {
        'page_obj': page_obj,
        'angajati': angajati,
        'angajat_selectat': angajat_id,
        'active_menu': 'pontaje',
    }
    return render(request, 'core/hr_pontaje.html', context)


@hr_required
def hr_concedii(request):
    status_filter = request.GET.get('status', 'in_asteptare')
    tip_filter = request.GET.get('tip', 'toate')
    cereri_qs = CerereConcediu.objects.select_related('angajat').order_by('-data_creare')
    if status_filter != 'toate':
        cereri_qs = cereri_qs.filter(status=status_filter)
    if tip_filter != 'toate':
        cereri_qs = cereri_qs.filter(tip_concediu=tip_filter)
    paginator = Paginator(cereri_qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    context = {'page_obj': page_obj, 'status_filter': status_filter, 'tip_filter': tip_filter, 'active_menu': 'concedii'}
    return render(request, 'core/hr_concedii.html', context)


@hr_required
@require_POST
def hr_concediu_actiune(request, cerere_id, actiune):
    cerere = get_object_or_404(CerereConcediu, pk=cerere_id)
    if actiune == 'aproba':
        cerere.status = 'aprobat'
        cerere.save()
        messages.success(request, f"Cererea lui {cerere.angajat.username} a fost aprobată.")
    elif actiune == 'respinge':
        cerere.status = 'respins'
        cerere.save()
        messages.warning(request, f"Cererea lui {cerere.angajat.username} a fost respinsă.")
    return redirect('hr_concedii')


@login_required
@require_POST
def cerere_concediu_creare(request):
    tip_concediu = request.POST.get('tip_concediu', 'odihna')
    data_inceput = request.POST.get('data_inceput')
    data_sfarsit = request.POST.get('data_sfarsit')
    motiv = request.POST.get('motiv', '')

    if not data_inceput or not data_sfarsit:
        messages.error(request, "Trebuie să completezi datele de început și sfârșit.")
        return redirect('dashboard')

    CerereConcediu.objects.create(
        angajat=request.user,
        tip_concediu=tip_concediu,
        data_inceput=data_inceput,
        data_sfarsit=data_sfarsit,
        motiv=motiv,
    )

    if tip_concediu == 'medical' and request.FILES.get('document_medical'):
        DocumentAngajat.objects.create(
            angajat=request.user,
            titlu=f"Concediu Medical {data_inceput} - {data_sfarsit}",
            fisier=request.FILES['document_medical'],
        )

    messages.success(request, "Cererea de concediu a fost trimisă.")
    return redirect('dashboard')


@login_required
@require_POST
def editare_profil(request):
    profil, _ = ProfileAngajat.objects.get_or_create(
        user=request.user,
        defaults={'functia': 'Șofer', 'departament': 'Transport', 'data_angajare': timezone.now().date()}
    )
    functia = request.POST.get('functia', '').strip()
    telefon = request.POST.get('telefon', '').strip()

    if functia:
        profil.functia = functia
    profil.telefon = telefon
    profil.save()

    messages.success(request, "Profilul a fost actualizat.")
    return redirect('dashboard')


@hr_required
def hr_angajat_detaliu(request, user_id):
    angajat_user = get_object_or_404(User, pk=user_id)
    profil, _ = ProfileAngajat.objects.get_or_create(
        user=angajat_user,
        defaults={'functia': 'Șofer', 'departament': 'Transport', 'data_angajare': timezone.now().date()}
    )

    an_curent = timezone.localtime(timezone.now()).year
    concedii_odihna_aprobate = CerereConcediu.objects.filter(
        angajat=angajat_user,
        tip_concediu='odihna',
        status='aprobat',
        data_inceput__year=an_curent,
    )
    zile_folosite = 0
    for c in concedii_odihna_aprobate:
        zile_folosite += (c.data_sfarsit - c.data_inceput).days + 1
    zile_ramase = max(profil.zile_concediu_alocate - zile_folosite, 0)

    if request.method == 'POST':
        profil.functia = request.POST.get('functia', profil.functia).strip()
        profil.departament = request.POST.get('departament', profil.departament).strip()
        profil.telefon = request.POST.get('telefon', '').strip()
        profil.cnp = request.POST.get('cnp', '').strip()
        profil.adresa = request.POST.get('adresa', '').strip()
        profil.tip_contract = request.POST.get('tip_contract', profil.tip_contract)

        salariu_baza = request.POST.get('salariu_baza', '0')
        try:
            profil.salariu_baza = float(salariu_baza)
        except (TypeError, ValueError):
            pass

        data_angajare = request.POST.get('data_angajare')
        if data_angajare:
            profil.data_angajare = data_angajare

        zile_alocate = request.POST.get('zile_concediu_alocate', '20')
        try:
            profil.zile_concediu_alocate = int(zile_alocate)
        except (TypeError, ValueError):
            pass

        profil.save()
        messages.success(request, f"Fișa lui {angajat_user.username} a fost actualizată.")
        return redirect('hr_angajat_detaliu', user_id=user_id)

    context = {
        'angajat_user': angajat_user,
        'profil': profil,
        'zile_folosite': zile_folosite,
        'zile_ramase': zile_ramase,
        'active_menu': 'angajati',
    }
    return render(request, 'core/hr_angajat_detaliu.html', context)


@hr_required
@require_POST
def hr_angajat_toggle_activ(request, user_id):
    angajat_user = get_object_or_404(User, pk=user_id)
    profil, _ = ProfileAngajat.objects.get_or_create(
        user=angajat_user,
        defaults={'functia': 'Șofer', 'departament': 'Transport', 'data_angajare': timezone.now().date()}
    )

    if profil.activ:
        profil.activ = False
        profil.data_incetare = timezone.now().date()
        messages.warning(request, f"{angajat_user.username} a fost marcat ca inactiv.")
    else:
        profil.activ = True
        profil.data_incetare = None
        messages.success(request, f"{angajat_user.username} a fost reactivat.")

    profil.save()
    return redirect('hr_angajat_detaliu', user_id=user_id)


@hr_required
def hr_calendar_concedii(request):
    import calendar as calendar_lib

    azi = timezone.localtime(timezone.now()).date()
    an = int(request.GET.get('an', azi.year))
    luna = int(request.GET.get('luna', azi.month))

    prima_zi = azi.replace(year=an, month=luna, day=1)
    ultima_zi_nr = calendar_lib.monthrange(an, luna)[1]
    ultima_zi = prima_zi.replace(day=ultima_zi_nr)

    cereri = CerereConcediu.objects.select_related('angajat').filter(
        status='aprobat',
        data_inceput__lte=ultima_zi,
        data_sfarsit__gte=prima_zi,
    )

    zile_cu_concedii = {}
    for c in cereri:
        start = max(c.data_inceput, prima_zi)
        stop = min(c.data_sfarsit, ultima_zi)
        zi_curenta = start
        while zi_curenta <= stop:
            nume = c.angajat.get_full_name() or c.angajat.username
            zile_cu_concedii.setdefault(zi_curenta, []).append(f"{nume} ({c.get_tip_concediu_display()})")
            zi_curenta += timezone.timedelta(days=1)

    zile_sortate = sorted(zile_cu_concedii.items())

    luna_precedenta = luna - 1
    an_precedent = an
    if luna_precedenta < 1:
        luna_precedenta = 12
        an_precedent -= 1

    luna_urmatoare = luna + 1
    an_urmator = an
    if luna_urmatoare > 12:
        luna_urmatoare = 1
        an_urmator += 1

    nume_luni = [
        'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
        'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie',
    ]

    context = {
        'zile_sortate': zile_sortate,
        'luna_nume': nume_luni[luna - 1],
        'an': an,
        'luna': luna,
        'an_precedent': an_precedent,
        'luna_precedenta': luna_precedenta,
        'an_urmator': an_urmator,
        'luna_urmatoare': luna_urmatoare,
        'active_menu': 'calendar',
    }
    return render(request, 'core/hr_calendar.html', context)


@hr_required
def hr_export_lunar(request):
    azi = timezone.localtime(timezone.now()).date()
    an = int(request.GET.get('an', azi.year))
    luna = int(request.GET.get('luna', azi.month))

    nume_luni = [
        'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
        'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie',
    ]

    angajati = ProfileAngajat.objects.select_related('user').filter(rol='angajat', activ=True).order_by('user__username')

    randuri = []
    for profil in angajati:
        pontaje_luna = Pontaj.objects.filter(
            angajat=profil.user,
            data__year=an,
            data__month=luna,
        )
        total_ore = sum(p.ore_lucrate for p in pontaje_luna)
        total_zile = pontaje_luna.count()
        salariu_baza = float(profil.salariu_baza or 0)

        randuri.append({
            'nume': profil.user.get_full_name() or profil.user.username,
            'functia': profil.functia,
            'salariu_baza': salariu_baza,
            'total_zile_lucrate': total_zile,
            'total_ore_lucrate': round(total_ore, 2),
        })

    context = {
        'randuri': randuri,
        'an': an,
        'luna': luna,
        'luna_nume': nume_luni[luna - 1],
        'active_menu': 'export',
    }
    return render(request, 'core/hr_export.html', context)


@hr_required
def hr_export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    azi = timezone.localtime(timezone.now()).date()
    an = int(request.GET.get('an', azi.year))
    luna = int(request.GET.get('luna', azi.month))

    angajati = ProfileAngajat.objects.select_related('user').filter(rol='angajat', activ=True).order_by('user__username')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Salarizare {luna}-{an}"

    headere = ['Nume', 'Funcție', 'Salariu Bază', 'Zile Lucrate', 'Ore Lucrate']
    ws.append(headere)
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')

    for profil in angajati:
        pontaje_luna = Pontaj.objects.filter(angajat=profil.user, data__year=an, data__month=luna)
        total_ore = sum(p.ore_lucrate for p in pontaje_luna)
        total_zile = pontaje_luna.count()
        salariu_baza = float(profil.salariu_baza or 0)

        ws.append([
            profil.user.get_full_name() or profil.user.username,
            profil.functia,
            salariu_baza,
            total_zile,
            round(total_ore, 2),
        ])

    for coloana in ws.columns:
        lungime_max = max(len(str(celula.value)) for celula in coloana)
        ws.column_dimensions[coloana[0].column_letter].width = lungime_max + 3

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="salarizare_{luna}_{an}.xlsx"'
    wb.save(response)
    return response


@hr_required
def hr_angajat_creare(request):
    parola_generata = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        functia = request.POST.get('functia', 'Șofer').strip()
        departament = request.POST.get('departament', 'Transport').strip()
        telefon = request.POST.get('telefon', '').strip()

        if not username:
            messages.error(request, "Username-ul este obligatoriu.")
            return redirect('hr_angajat_creare')

        if User.objects.filter(username=username).exists():
            messages.error(request, f"Username-ul '{username}' este deja folosit.")
            return redirect('hr_angajat_creare')

        from django.utils.crypto import get_random_string
        parola_generata = get_random_string(10)

        user_nou = User.objects.create_user(
            username=username,
            password=parola_generata,
            first_name=first_name,
            last_name=last_name,
        )
        ProfileAngajat.objects.create(
            user=user_nou,
            functia=functia,
            departament=departament,
            telefon=telefon,
            data_angajare=timezone.now().date(),
        )
        messages.success(request, f"Contul pentru {username} a fost creat.")

        context = {'parola_generata': parola_generata, 'username_creat': username, 'active_menu': 'angajati'}
        return render(request, 'core/hr_angajat_creare.html', context)

    context = {'active_menu': 'angajati'}
    return render(request, 'core/hr_angajat_creare.html', context)


@login_required
@require_POST
def schimbare_parola(request):
    from django.contrib.auth import update_session_auth_hash

    parola_curenta = request.POST.get('parola_curenta', '')
    parola_noua = request.POST.get('parola_noua', '')
    parola_noua_confirm = request.POST.get('parola_noua_confirm', '')

    if not request.user.check_password(parola_curenta):
        messages.error(request, "Parola curentă este greșită.")
        return redirect('dashboard')

    if len(parola_noua) < 8:
        messages.error(request, "Parola nouă trebuie să aibă minimum 8 caractere.")
        return redirect('dashboard')

    if parola_noua != parola_noua_confirm:
        messages.error(request, "Parolele noi nu coincid.")
        return redirect('dashboard')

    request.user.set_password(parola_noua)
    request.user.save()
    update_session_auth_hash(request, request.user)
    messages.success(request, "Parola a fost schimbată cu succes.")
    return redirect('dashboard')